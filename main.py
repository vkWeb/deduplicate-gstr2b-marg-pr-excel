"""
      Col Name                   Index       
-------------------          ------------
GSTIN of supplier               0 / A	
Trade/Legal name                1 / B
Type of inward supplies         2 / C 
Document type                   3 / D
Document number                 4 / E
Document date                   5 / F
Taxable value (₹)               6 / G 
Integrated tax (₹)              7 / H
Central tax (₹)                 8 / I
State/ UT tax (₹)               9 / J
Cess (₹)                        10 / K
"""
from openpyxl import load_workbook


MARG_GSTR2A_PURCHASE_REGISTER = "22aezpa0778g1zf_purch_register_202209.xlsx"
PURCHASE_REGISTER_SHEET_NAME = "Purchase Reigster"
PURCHASE_REGISTER_MAP = {}


wb = load_workbook(MARG_GSTR2A_PURCHASE_REGISTER)
ws = wb[PURCHASE_REGISTER_SHEET_NAME]


def build_pr_map():
    for row in ws.iter_rows(min_row=6, max_col=11, values_only=True):
        row = list(row)
        if row[0] is None:
            break

        try:
            key = row[0] + "_" + row[4]
            existing_row = PURCHASE_REGISTER_MAP[key]
            for i in range(6, 10 + 1):
                existing_row[i] = existing_row[i] + row[i]
            PURCHASE_REGISTER_MAP[key] = existing_row
        except KeyError:
            PURCHASE_REGISTER_MAP[key] = row


def reset_all_rows():
    for row in ws.iter_rows(min_row=6, max_col=11):
        if row[0] is None:
            break

        for col in row:
            col.value = ""


def write_pr_map():
    row_number = 6
    for key, value in PURCHASE_REGISTER_MAP.items():

        if not len(value) == 11:
            raise Exception("Purchase Register Hashmap is corrupted :(")

        for i in range(0, len(value)):
            ws.cell(row=row_number, column=i+1, value=value[i])

        row_number = row_number + 1


if __name__ == "__main__":
    build_pr_map()
    reset_all_rows()
    write_pr_map()
    wb.save("updated.xlsx")
